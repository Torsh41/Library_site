from openpyxl import load_workbook
from copy import copy


def sort_with_mask(t):
    mask = {'Обложка': 0, 'Название': 1, 'ISBN': 2, 'Автор': 3, 'Издательство': 4, 'Описание': 5, 'Дата выхода': 6, 'Количество глав': 7, 'Категория': 8}
    # Сортируем по значению в словаре mask
    return mask[t[0]]    


def parse_excel(file) -> list | bool:
    HEADERS_COUNT = 11
    HEADERS = {'Название', 'Авторы', 'Серия', 'Категории', 'Дата публикации', 'Издательство', 'Количество страниц', 'ISBN', 'Комментарии', 'Краткое содержание', 'Ссылка'}       
    wb = load_workbook(file)
    sheet = wb[wb.active.title]
    data = list(); columns_count = sheet.max_column
    if columns_count != HEADERS_COUNT:
        return False

    is_header_ord = True; headers_set = set()
    for row in sheet:  
        # проверяем на соответствие нашим ожиданиям входной файл 
        if is_header_ord:
            for index in range(HEADERS_COUNT):  
                headers_set.add(str(row[index].value).strip())
            if headers_set != HEADERS:
                return False
            is_header_ord = False
            continue
        
        cur_book = list(); count = 1  
        for index in range(len(row)):  
            if (index == 8 and row[index].value != "Комментарии"):  
                val = str(row[index].value).split(",")[-1]
                if "экземпляра" in val or "экземпляров" in val or "экземпляр" in val:  
                    count = int(val.strip().split(" ")[0])
                     
            val = row[index].value
            if type(val) == int or type(val) == float:
                cur_book.append(val)
            elif type(val) == str:
                cur_book.append(str(val).strip().lower())
            else:
                cur_book.append(str("неизвестно"))
           
        cur_book.append(count)    
        data.append(copy(cur_book)) 
    
    del data[0]
    return data
 
 
def add_many_books(file) -> list | bool:
    HEADERS_COUNT = 9
    HEADERS = ['Обложка', 'Название', 'ISBN', 'Автор', 'Издательство', 'Описание', 'Дата выхода', 'Количество глав', 'Категория']
    wb = load_workbook(file)
    sheet = wb[wb.active.title]
    books = dict(); headers = dict(); columns_count = sheet.max_column; max_row = sheet.max_row - 1
    columns = [list() for _ in range(columns_count)]
    if columns_count != HEADERS_COUNT:
        return False
    is_header_ord = True
    for row in sheet:  
        # проверяем на соответствие нашим ожиданиям входной файл 
        if is_header_ord:
            count = 0
            for index in range(columns_count): 
                headers[str(row[index].value).strip()] = count
                count += 1
            is_header_ord = False
            if list(dict(sorted(headers.items(), key=sort_with_mask)).keys()) != HEADERS:
                return False
            continue
        
        for index in range(len(row)):  
            if headers.get(HEADERS[0]) == index or headers.get(HEADERS[1]) == index or headers.get(HEADERS[2]) == index or headers.get(HEADERS[3]) == index \
            or headers.get(HEADERS[4]) == index or headers.get(HEADERS[5]) == index or headers.get(HEADERS[8]) == index or headers.get(HEADERS[6]) == index:
                elem = str(row[index].value).strip().lower()
                if elem:
                    columns[index].append(elem)
                else:
                    columns[index].append(None)
            elif headers.get(HEADERS[7]) == index:
                elem = int(row[index].value)
                if elem:
                    columns[index].append(elem)
                else:
                    columns[index].append(None)
               
    for header, column in zip(headers.keys(), columns):
        books[header] = column
        
    books = dict(sorted(books.items(), key=sort_with_mask))
    books_elems = list(books.values()); books = list()
    for book_elem_index in range(max_row):
        books.append([books_elems[0][book_elem_index], books_elems[1][book_elem_index], books_elems[2][book_elem_index], 
                      books_elems[3][book_elem_index], books_elems[4][book_elem_index], books_elems[5][book_elem_index],
                      books_elems[6][book_elem_index], books_elems[7][book_elem_index], books_elems[8][book_elem_index]])

    return books
    